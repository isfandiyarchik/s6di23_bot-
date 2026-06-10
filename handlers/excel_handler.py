import io
import os
import math
import logging
from database import db_cursor, now_uz
from handlers.common import (
    is_admin, check_access,
    parse_birth_date, clean_hemis,
    back_menu, excel_submenu
)

logger = logging.getLogger(__name__)

def register(bot):
    ca = check_access(bot)

    @bot.message_handler(func=lambda m: m.text == "📊 Excel басқарыу")
    @ca
    def excel_management(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        bot.send_message(message.chat.id, "📊 <b>Excel басқарыу</b>", reply_markup=excel_submenu())

    @bot.message_handler(func=lambda m: m.text == "📥 Excel жүклеу")
    @ca
    def excel_download(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        _excel_download_impl(bot, message)

    @bot.message_handler(func=lambda m: m.text == "📤 Excel импорт")
    @ca
    def excel_import_start(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        msg = bot.send_message(message.chat.id,
            "📤 <b>Excel файлды жибериңиз (.xlsx):</b>", reply_markup=back_menu())
        bot.register_next_step_handler(msg, lambda m: handle_excel_import(bot, m))

def handle_excel_import(bot, message):
    if not is_admin(message.from_user.id): return
    if message.text and message.text == "⬅️ Артқа":
        bot.send_message(message.chat.id, "📊 Excel басқарыу", reply_markup=excel_submenu()); return
    _excel_import_impl(bot, message)

def _excel_download_impl(bot, message):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        bot.send_message(message.chat.id, "❌ openpyxl орнатылмаған!", reply_markup=excel_submenu()); return
    try:
        with db_cursor() as (_, cursor):
            cursor.execute("SELECT id,full_name,birth_date,phone,hemis,username FROM students ORDER BY full_name")
            rows = cursor.fetchall()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Студентлер"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="2E75B6")
        ca = Alignment(horizontal="center", vertical="center")
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        alt = PatternFill("solid", fgColor="EBF3FB")
        headers = ["№","ФИО","Тууылған күни","Телефон","HEMIS","Telegram","TelegramID"]
        widths = [5,38,16,20,18,22,16]
        for ci,(h,w) in enumerate(zip(headers,widths),1):
            cell = ws.cell(row=1,column=ci,value=h)
            cell.font=hf; cell.fill=hfill; cell.alignment=ca; cell.border=tb
            ws.column_dimensions[cell.column_letter].width=w
        ws.row_dimensions[1].height=22
        for i,row in enumerate(rows,1):
            tg_id,full_name,birth_date,phone,hemis,username=row
            hemis_val=clean_hemis(hemis); uname=f"@{username}" if username else ""
            rfill=alt if i%2==0 else None
            for ci,val in enumerate([i,full_name or "",birth_date or "",phone or "",hemis_val,uname,tg_id],1):
                cell=ws.cell(row=i+1,column=ci,value=val)
                cell.border=tb; cell.alignment=ca
                if rfill: cell.fill=rfill
        for r in range(2,len(rows)+2):
            ws.cell(row=r,column=7).number_format='0'
            cell_bd=ws.cell(row=r,column=3)
            if cell_bd.value:
                for fmt in ("%Y-%m-%d","%d.%m.%Y","%d/%m/%Y","%Y/%m/%d"):
                    try:
                        from datetime import datetime
                        cell_bd.value=datetime.strptime(str(cell_bd.value),fmt)
                        cell_bd.number_format='DD.MM.YYYY'; break
                    except: pass
        ws.auto_filter.ref=f"A1:G{len(rows)+1}"; ws.freeze_panes="A2"
        path="/tmp/students_export.xlsx"; wb.save(path)
        with open(path,"rb") as f: data=f.read()
        try: os.remove(path)
        except: pass
        fo=io.BytesIO(data); fo.name="students_export.xlsx"
        bot.send_document(message.chat.id, fo, caption=(
            f"📥 <b>Студентлер дизими</b> — {len(rows)} студент\n\n"
            "⚠️ <b>G қатарын (TelegramID) өзгертпеңиз!</b>\n"
            "📤 Толтырып болғаннан кейин <b>Excel импорт</b> арқалы жүклеңиз."))
        bot.send_message(message.chat.id, "✅ Жиберилди!", reply_markup=excel_submenu())
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Excel жасауда қате: {e}", reply_markup=excel_submenu())

def _excel_import_impl(bot, message):
    if not message.document:
        msg=bot.send_message(message.chat.id,"⚠️ .xlsx файл жибериңиз!",reply_markup=back_menu())
        bot.register_next_step_handler(msg,lambda m: handle_excel_import(bot,m)); return
    fname=message.document.file_name or ""
    if not fname.lower().endswith(".xlsx"):
        msg=bot.send_message(message.chat.id,"⚠️ Тек .xlsx форматы қабылланады!",reply_markup=back_menu())
        bot.register_next_step_handler(msg,lambda m: handle_excel_import(bot,m)); return
    try:
        import openpyxl
    except ImportError:
        bot.send_message(message.chat.id,"❌ openpyxl орнатылмаған!",reply_markup=excel_submenu()); return
    path="/tmp/import_students.xlsx"
    try:
        fi=bot.get_file(message.document.file_id)
        with open(path,"wb") as f: f.write(bot.download_file(fi.file_path))
    except Exception as e:
        bot.send_message(message.chat.id,f"❌ Файлды жүклеу мүмкин болмады: {e}",reply_markup=excel_submenu()); return
    try:
        wb=openpyxl.load_workbook(path,data_only=True); ws=wb.active
    except Exception as e:
        bot.send_message(message.chat.id,f"❌ Excel файлды оқыу мүмкин болмады: {e}",reply_markup=excel_submenu())
        try: os.remove(path)
        except: pass
        return

    def clean_cell(val):
        if val is None: return ""
        if isinstance(val,float) and math.isnan(val): return ""
        s=str(val).strip()
        if s in ("None","nan",""): return ""
        if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): return s[:-2]
        return s

    def parse_tg_id(val):
        if val is None: return None
        try:
            if isinstance(val,float):
                if math.isnan(val): return None
                return int(val)
            if isinstance(val,int): return val
            s=str(val).strip().split(".")[0]
            return int(s) if s.lstrip("-").isdigit() else None
        except: return None

    updated=added=skipped=errors=0
    try:
        with db_cursor() as (conn,cursor):
            for row_idx,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
                try:
                    if not row or all(v is None for v in row): continue
                    full_name=clean_cell(row[1] if len(row)>1 else None)
                    birth_date=parse_birth_date(row[2] if len(row)>2 else None)
                    phone=clean_cell(row[3] if len(row)>3 else None)
                    hemis=clean_cell(row[4] if len(row)>4 else None)
                    uname_raw=clean_cell(row[5] if len(row)>5 else None)
                    tg_id_raw=row[6] if len(row)>6 else None
                    if not full_name or full_name=="ФИО": skipped+=1; continue
                    uname=uname_raw.lstrip("@") if uname_raw else None
                    if not uname: uname=None
                    tg_id=parse_tg_id(tg_id_raw)
                    if tg_id:
                        cursor.execute("SELECT id FROM students WHERE id=%s",(tg_id,))
                        if cursor.fetchone():
                            if uname:
                                cursor.execute(
                                    "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s,username=%s WHERE id=%s",
                                    (full_name,birth_date,phone,hemis,uname,tg_id))
                            else:
                                cursor.execute(
                                    "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s WHERE id=%s",
                                    (full_name,birth_date,phone,hemis,tg_id))
                            updated+=1
                        else:
                            cursor.execute(
                                "INSERT INTO students(id,username,last_active,full_name,birth_date,phone,hemis,started) "
                                "VALUES(%s,%s,%s,%s,%s,%s,%s,0)",
                                (tg_id,uname,now_uz(),full_name,birth_date,phone,hemis))
                            added+=1
                        continue
                    if uname:
                        cursor.execute("SELECT id FROM students WHERE username=%s",(uname,))
                        if cursor.fetchone():
                            cursor.execute(
                                "UPDATE students SET full_name=%s,birth_date=%s,phone=%s,hemis=%s WHERE username=%s",
                                (full_name,birth_date,phone,hemis,uname))
                            updated+=1; continue
                    cursor.execute("SELECT id FROM students WHERE full_name=%s",(full_name,))
                    if cursor.fetchone():
                        cursor.execute("UPDATE students SET birth_date=%s,phone=%s,hemis=%s WHERE full_name=%s",
                            (birth_date,phone,hemis,full_name))
                        updated+=1
                    else: skipped+=1
                except Exception as e:
                    logger.warning(f"Import row {row_idx}: {e}")
                    try: conn.rollback()
                    except: pass
                    errors+=1; continue
            conn.commit()
    except Exception as e:
        bot.send_message(message.chat.id,f"❌ Импортта критикалық қате: {e}",reply_markup=excel_submenu()); return
    finally:
        try: os.remove(path)
        except: pass
    bot.send_message(message.chat.id,
        f"✅ <b>Импорт жуумақланды!</b>\n\n"
        f"🔄 Тазаланды: <b>{updated}</b>\n"
        f"➕ Қосылды:   <b>{added}</b>\n"
        f"⏭ Өткизилди: <b>{skipped}</b>\n"
        f"❌ Қателер:   <b>{errors}</b>",
        reply_markup=excel_submenu())

