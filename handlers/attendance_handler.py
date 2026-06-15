# Attendance handler — барлау + студент өз барлауын көреді + статистика
import io
import re
import os
import logging
from telebot import types
from database import (db_cursor, now_uz, save_attendance_session,
    load_attendance_session, delete_attendance_session)
from handlers.common import (
    is_admin, check_access, check_access_cb,
    DAYS_EN_TO_RU, MONTHS_RU,
    date_to_ru, attendance_submenu, main_menu, back_menu
)

logger = logging.getLogger(__name__)


def generate_attendance_excel(students, results, date_str, para, subject):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None
    try:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Барлау"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="2E75B6")
        ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
        la = Alignment(horizontal="left", vertical="center", wrap_text=True)
        gf = PatternFill("solid", fgColor="C6EFCE")
        rf = PatternFill("solid", fgColor="FFC7CE")
        tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
        headers = ["№", "ФИО", "Күн", "Пара", "Пән", "Барлау"]
        widths = [5, 38, 14, 8, 22, 12]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hf; cell.fill = hfill
            cell.alignment = ca; cell.border = tb
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 20
        rn = 2
        for i, item in enumerate(students, 1):
            if not isinstance(item, (list, tuple)) or len(item) < 2: continue
            first, second = item[0], item[1]
            if isinstance(first, int):
                name = str(second) if second else "—"
                status = results.get(first, "absent") if results else "absent"
            else:
                name = str(first) if first else "—"
                status = str(second) if second else "absent"
            st_text = "✅ Бар" if status == "present" else "❌ Жоқ"
            rfill = gf if status == "present" else rf
            for ci, val in enumerate([i, name, date_str, para, subject, st_text], 1):
                cell = ws.cell(row=rn, column=ci, value=val)
                cell.fill = rfill; cell.border = tb
                cell.alignment = la if ci == 2 else ca
            rn += 1
        ws.auto_filter.ref = f"A1:F{rn-1}"
        safe_date = date_str.replace("-", "")
        safe_subj = re.sub(r'[^\w]', '_', subject)[:20]
        path = f"/tmp/attendance_{safe_date}_para{para}_{safe_subj}.xlsx"
        wb.save(path); return path
    except Exception as e:
        logger.error(f"generate_attendance_excel: {e}"); return None


def send_excel_file(bot, chat_id, path, caption=""):
    try:
        with open(path, "rb") as f: data = f.read()
        try: os.remove(path)
        except: pass
        fo = io.BytesIO(data); fo.name = os.path.basename(path)
        bot.send_document(chat_id, fo, caption=caption); return True
    except Exception as e:
        logger.error(f"send_excel_file: {e}"); return False


def build_attendance_markup(session):
    idx = session["current_index"]
    if idx >= len(session["students"]): return None, None
    student = session["students"][idx]
    sid, sname = student[0], student[1]
    total = len(session["students"]); done = len(session["results"])
    markup = types.InlineKeyboardMarkup()
    markup.row(
        types.InlineKeyboardButton("✅ Бар", callback_data=f"att_mark_present_{sid}"),
        types.InlineKeyboardButton("❌ Жоқ", callback_data=f"att_mark_absent_{sid}"))
    markup.add(types.InlineKeyboardButton("🏁 Жуумақлау", callback_data="att_finish"))
    sep = "─" * 30
    text = (f"📊 <b>Барлау — {session['para']}-пара: {session['subject']}</b>\n"
            f"📅 {session['date']}\n{sep}\n👤 <b>{sname}</b>\n{sep}\n"
            f"<i>{done}/{total} белгиленди</i>")
    return text, markup


def finish_attendance(bot, message, admin_id):
    session = load_attendance_session(admin_id)
    if not session: return
    delete_attendance_session(admin_id)
    date_str = session["date"]; para = session["para"]
    subject = session["subject"]; students = session["students"]
    results = session["results"]
    present_list = []; absent_list = []
    try:
        with db_cursor() as (conn, cursor):
            for item in students:
                sid, sname = item[0], item[1]
                status = results.get(sid, "absent")
                cursor.execute(
                    "INSERT INTO attendance(date,para,subject,student_id,student_name,status) "
                    "VALUES(%s,%s,%s,%s,%s,%s)",
                    (date_str, para, subject, sid, sname, status))
                if status == "present": present_list.append(sname)
                else: absent_list.append((sid, sname))
            conn.commit()
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Барлауды сақлауда қате: {e}"); return
    total = len(students)
    sep = "─" * 30
    result_text = (
        f"📊 <b>Барлау нәтийжеси сақланды!</b>\n"
        f"📅 {date_str} | {para}-пара: <b>{subject}</b>\n{sep}\n"
        f"✅ Бар: <b>{len(present_list)}/{total}</b>\n"
        f"❌ Жоқ: <b>{len(absent_list)}/{total}</b>\n{sep}\n")
    if absent_list:
        result_text += "❌ <b>Жоқлар:</b>\n"
        for _, n in absent_list: result_text += f"  • {n}\n"
    else:
        result_text += "🎉 Барлық студентлер бар!\n"
    try:
        bot.edit_message_text(result_text, message.chat.id, message.message_id,
            parse_mode="HTML", reply_markup=None)
    except:
        bot.send_message(message.chat.id, result_text, parse_mode="HTML")
    path = generate_attendance_excel(students, results, date_str, para, subject)
    if path:
        send_excel_file(bot, message.chat.id, path,
            caption=f"📊 {date_str} | {para}-пара: {subject}")
    for sid, sname in absent_list:
        try:
            bot.send_message(sid,
                f"⚠️ <b>Ескертиу!</b>\n\nСиз бүгин <b>{para}-парада</b> "
                f"(<b>{subject}</b>) болмадыңыз!\n"
                f"📅 {date_str}\n\nСебебиңизди группаға хабарлаңыз.")
        except: pass
    bot.send_message(message.chat.id,
        "✅ Барлау сақланды!\n📅 Тарийхты <b>Барлау тарийхы</b> арқалы ашыңыз.",
        reply_markup=attendance_submenu())


def register(bot):
    ca = check_access(bot)
    cacb = check_access_cb(bot)

    @bot.message_handler(func=lambda m: m.text == "📊 Барлау басқарыу")
    @ca
    def attendance_management(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        bot.send_message(message.chat.id, "📊 <b>Барлау басқарыу</b>",
            reply_markup=attendance_submenu())

    @bot.message_handler(func=lambda m: m.text == "📊 Барлау")
    @ca
    def start_attendance(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        today = DAYS_EN_TO_RU.get(now_uz().strftime("%A"), "")
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT subject,time FROM schedule WHERE day=%s ORDER BY time", (today,))
            lessons = cursor.fetchall()
        if not lessons:
            bot.send_message(message.chat.id,
                f"📭 Бүгин ({today}) сабақ жоқ.",
                reply_markup=attendance_submenu()); return
        markup = types.InlineKeyboardMarkup()
        for i, (subject, time_) in enumerate(lessons, 1):
            markup.add(types.InlineKeyboardButton(
                text=f"{i}-пара: {subject} ({time_})",
                callback_data=f"att_para_{i}"))
        bot.send_message(message.chat.id,
            f"📊 <b>Барлау — {today}</b>\n\nҚай параны белгилейсиз:",
            reply_markup=markup)

    @bot.message_handler(func=lambda m: m.text == "📅 Барлау тарийхы")
    @ca
    def attendance_history(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT DISTINCT LEFT(date,7) as ym FROM attendance ORDER BY ym DESC")
            months = [r[0] for r in cursor.fetchall()]
        if not months:
            bot.send_message(message.chat.id, "📭 Барлау жазылмаған.",
                reply_markup=attendance_submenu()); return
        markup = types.InlineKeyboardMarkup()
        for ym in months:
            y, mo = ym.split("-")
            markup.add(types.InlineKeyboardButton(
                text=f"📅 {MONTHS_RU.get(int(mo), mo)} {y}",
                callback_data=f"hist_month_{ym}"))
        bot.send_message(message.chat.id,
            "📅 <b>Барлау тарийхы</b>\n\nАйды таңлаңыз:", reply_markup=markup)

    # ── ЖАҢ: Студент өз барлауын көреді ─────────────────────
    @bot.message_handler(func=lambda m: m.text == "📊 Мениң барлауым")
    @ca
    def my_attendance(message):
        uid = message.from_user.id
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT full_name FROM students WHERE id=%s", (uid,))
            row = cursor.fetchone()
            if not row:
                bot.send_message(message.chat.id, "❌ Студент табылмады.",
                    reply_markup=main_menu(uid)); return
            sname = row[0]
            cursor.execute(
                "SELECT date,para,subject,status FROM attendance "
                "WHERE student_id=%s ORDER BY date DESC, para ASC LIMIT 30",
                (uid,))
            records = cursor.fetchall()
            cursor.execute(
                "SELECT COUNT(*) FROM attendance WHERE student_id=%s", (uid,))
            total = cursor.fetchone()[0]
            cursor.execute(
                "SELECT COUNT(*) FROM attendance WHERE student_id=%s AND status='present'",
                (uid,))
            present = cursor.fetchone()[0]

        if not records:
            bot.send_message(message.chat.id,
                "📭 Барлау жазбаңыз жоқ.", reply_markup=main_menu(uid)); return

        absent = total - present
        pct = int((present / total) * 100) if total > 0 else 0
        bar = "🟩" * (pct // 10) + "⬜" * (10 - pct // 10)
        sep = "─" * 30

        text = (f"📊 <b>Мениң барлауым</b>\n"
                f"👤 <b>{sname}</b>\n{sep}\n"
                f"✅ Барғаным:  <b>{present}</b>\n"
                f"❌ Болмаған: <b>{absent}</b>\n"
                f"📈 Улыума:    <b>{total}</b>\n"
                f"{bar} <b>{pct}%</b>\n{sep}\n\n"
                f"📋 <b>Соңғы 30 жазба:</b>\n{sep}\n")

        for rec in records:
            icon = "✅" if rec[3] == "present" else "❌"
            text += f"{icon} {date_to_ru(rec[0])} | {rec[1]}-пара: {rec[2]}\n"

        bot.send_message(message.chat.id, text, reply_markup=main_menu(uid))

    # ── ЖАҢ: Барлау статистикасы (admin) ────────────────────
    @bot.message_handler(func=lambda m: m.text == "📈 Барлау статистикасы")
    @ca
    def attendance_stats(message):
        if not is_admin(message.from_user.id):
            bot.send_message(message.chat.id, "🚫"); return
        with db_cursor() as (_, cursor):
            # Жалпы статистика
            cursor.execute("SELECT COUNT(DISTINCT date) FROM attendance")
            total_days = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM attendance WHERE status='present'")
            total_present = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM attendance")
            total_all = cursor.fetchone()[0]
            # Ең көп келмеген студенттер
            cursor.execute("""
                SELECT student_name, COUNT(*) as absent_count
                FROM attendance WHERE status='absent'
                GROUP BY student_name ORDER BY absent_count DESC LIMIT 5
            """)
            top_absent = cursor.fetchall()
            # Ең жиі келген студенттер
            cursor.execute("""
                SELECT student_name, COUNT(*) as present_count
                FROM attendance WHERE status='present'
                GROUP BY student_name ORDER BY present_count DESC LIMIT 5
            """)
            top_present = cursor.fetchall()
            # Пән бойынша статистика
            cursor.execute("""
                SELECT subject,
                    SUM(CASE WHEN status='present' THEN 1 ELSE 0 END) as p,
                    COUNT(*) as t
                FROM attendance GROUP BY subject ORDER BY subject
            """)
            subj_stats = cursor.fetchall()

        if total_all == 0:
            bot.send_message(message.chat.id, "📭 Барлау жазбасы жоқ.",
                reply_markup=attendance_submenu()); return

        pct = int((total_present / total_all) * 100) if total_all > 0 else 0
        bar = "🟩" * (pct // 10) + "⬜" * (10 - pct // 10)
        sep = "─" * 30
        sep2 = "═" * 30

        text = (f"📈 <b>БАРЛАУ СТАТИСТИКАСЫ</b>\n{sep2}\n\n"
                f"📅 Барлау күнлери: <b>{total_days}</b>\n"
                f"✅ Барғанлар:      <b>{total_present}</b>\n"
                f"❌ Болмағанлар:    <b>{total_all - total_present}</b>\n"
                f"📊 Орташа қатнасыу:  <b>{pct}%</b>\n"
                f"{bar}\n{sep}\n")

        if top_absent:
            text += "\n⚠️ <b>Ең көп болмағанлар:</b>\n"
            for i, (name, cnt) in enumerate(top_absent, 1):
                text += f"  {i}. {name} — <b>{cnt} рет</b>\n"

        if top_present:
            text += f"\n🏆 <b>Ең коп келгенлер:</b>\n"
            for i, (name, cnt) in enumerate(top_present, 1):
                text += f"  {i}. {name} — <b>{cnt} рет</b>\n"

        if subj_stats:
            text += f"\n{sep}\n📚 <b>Пән бойынша:</b>\n"
            for subj, p, t in subj_stats:
                s_pct = int((p / t) * 100) if t > 0 else 0
                text += f"  📖 {subj}: <b>{s_pct}%</b> ({p}/{t})\n"

        bot.send_message(message.chat.id, text, reply_markup=attendance_submenu())

    @bot.callback_query_handler(func=lambda c: c.data.startswith("att_para_"))
    @cacb
    def att_select_para(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "🚫 Тек admin-ге!"); return
        para = int(call.data.split("_")[2])
        date_str = now_uz().strftime("%Y-%m-%d")
        today = DAYS_EN_TO_RU.get(now_uz().strftime("%A"), "")
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT subject FROM schedule WHERE day=%s ORDER BY time", (today,))
            lessons = [r[0] for r in cursor.fetchall()]
        if para > len(lessons):
            bot.answer_callback_query(call.id, "Қате пара нөмири."); return
        subject = lessons[para - 1]
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT id,full_name FROM students "
                "WHERE full_name IS NOT NULL AND full_name!='' ORDER BY full_name")
            students = [[r[0], r[1]] for r in cursor.fetchall()]
        if not students:
            bot.answer_callback_query(call.id, "Студентлер дизими бос!")
            bot.edit_message_text("📭 Студентлерде ФИО жоқ.",
                call.message.chat.id, call.message.message_id); return
        session = {"date": date_str, "para": para, "subject": subject,
                   "students": students, "results": {}, "current_index": 0}
        save_attendance_session(call.from_user.id, session)
        text, markup = build_attendance_markup(session)
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id,
            reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("att_mark_"))
    @cacb
    def att_mark_student(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "🚫"); return
        session = load_attendance_session(call.from_user.id)
        if not session:
            bot.answer_callback_query(call.id, "Сессия табылмады."); return
        parts = call.data.split("_"); status = parts[2]
        try: sid = int(parts[3])
        except:
            bot.answer_callback_query(call.id, "Қате."); return
        session["results"][sid] = status
        session["current_index"] += 1
        save_attendance_session(call.from_user.id, session)
        sname = next((s[1] for s in session["students"] if s[0] == sid), "—")
        icon = "✅" if status == "present" else "❌"
        bot.answer_callback_query(call.id, f"{icon} {sname}")
        if session["current_index"] >= len(session["students"]):
            finish_attendance(bot, call.message, call.from_user.id)
        else:
            text, markup = build_attendance_markup(session)
            try:
                bot.edit_message_text(text, call.message.chat.id,
                    call.message.message_id, reply_markup=markup, parse_mode="HTML")
            except: pass

    @bot.callback_query_handler(func=lambda c: c.data == "att_finish")
    @cacb
    def att_finish_early(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id); return
        session = load_attendance_session(call.from_user.id)
        if not session:
            bot.answer_callback_query(call.id, "Сессия табылмады."); return
        bot.answer_callback_query(call.id, "Барлау жуумақланды!")
        finish_attendance(bot, call.message, call.from_user.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("hist_month_"))
    @cacb
    def hist_select_month(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "🚫"); return
        ym = call.data.replace("hist_month_", ""); y, mo = ym.split("-")
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT DISTINCT date FROM attendance "
                "WHERE LEFT(date,7)=%s ORDER BY date DESC", (ym,))
            days = [r[0] for r in cursor.fetchall()]
        if not days:
            bot.answer_callback_query(call.id, "Бұл айда барлау жоқ."); return
        markup = types.InlineKeyboardMarkup()
        for d in days:
            markup.add(types.InlineKeyboardButton(
                text=f"📆 {date_to_ru(d)}", callback_data=f"hist_day_{d}"))
        markup.add(types.InlineKeyboardButton("◀️ Назад", callback_data="hist_back_months"))
        bot.edit_message_text(
            f"📅 <b>{MONTHS_RU.get(int(mo), mo)} {y}</b>\n\nКүнди таңлаңыз:",
            call.message.chat.id, call.message.message_id,
            reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data == "hist_back_months")
    @cacb
    def hist_back_to_months(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id); return
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT DISTINCT LEFT(date,7) as ym FROM attendance ORDER BY ym DESC")
            months = [r[0] for r in cursor.fetchall()]
        markup = types.InlineKeyboardMarkup()
        for ym in months:
            y, mo = ym.split("-")
            markup.add(types.InlineKeyboardButton(
                text=f"📅 {MONTHS_RU.get(int(mo), mo)} {y}",
                callback_data=f"hist_month_{ym}"))
        bot.edit_message_text("📅 <b>Барлау тарийхы</b>\n\nАйды таңлаңыз:",
            call.message.chat.id, call.message.message_id,
            reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("hist_day_"))
    @cacb
    def hist_select_day(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "🚫"); return
        date_str = call.data.replace("hist_day_", ""); ym = date_str[:7]
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT DISTINCT para,subject FROM attendance "
                "WHERE date=%s ORDER BY para", (date_str,))
            paras = cursor.fetchall()
        if not paras:
            bot.answer_callback_query(call.id, "Бұл күнде барлау жоқ."); return
        markup = types.InlineKeyboardMarkup()
        for para, subject in paras:
            markup.add(types.InlineKeyboardButton(
                text=f"📖 {para}-пара: {subject}",
                callback_data=f"hist_para_{date_str}_{para}"))
        markup.add(types.InlineKeyboardButton(
            "◀️ Назад", callback_data=f"hist_month_{ym}"))
        bot.edit_message_text(
            f"📆 <b>{date_to_ru(date_str)}</b>\n\nПараны таңлаңыз:",
            call.message.chat.id, call.message.message_id,
            reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("hist_para_"))
    @cacb
    def hist_show_para(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "🚫"); return
        rest = call.data[len("hist_para_"):]; last_us = rest.rfind("_")
        date_str = rest[:last_us]; para = int(rest[last_us + 1:])
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT student_name,status FROM attendance "
                "WHERE date=%s AND para=%s ORDER BY student_name", (date_str, para))
            records = cursor.fetchall()
            cursor.execute(
                "SELECT DISTINCT subject FROM attendance WHERE date=%s AND para=%s",
                (date_str, para))
            sr = cursor.fetchone()
        subject = sr[0] if sr else "—"
        present = [r[0] for r in records if r[1] == "present"]
        absent = [r[0] for r in records if r[1] == "absent"]
        total = len(records)
        sep = "─" * 30
        text = (f"📊 <b>Барлау нәтийжеси</b>\n"
                f"📆 {date_to_ru(date_str)} | {para}-пара: <b>{subject}</b>\n{sep}\n"
                f"✅ Бар: <b>{len(present)}/{total}</b>\n"
                f"❌ Жоқ: <b>{len(absent)}/{total}</b>\n{sep}\n")
        if present:
            text += "✅ <b>Барлар:</b>\n" + "".join(f"  • {n}\n" for n in present) + "\n"
        if absent:
            text += "❌ <b>Жоқлар:</b>\n" + "".join(f"  • {n}\n" for n in absent)
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton(
            "📥 Excel жүклеу", callback_data=f"hist_excel_{date_str}_{para}"))
        markup.add(types.InlineKeyboardButton(
            "◀️ Назад", callback_data=f"hist_day_{date_str}"))
        bot.edit_message_text(text, call.message.chat.id, call.message.message_id,
            reply_markup=markup, parse_mode="HTML")
        bot.answer_callback_query(call.id)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("hist_excel_"))
    @cacb
    def hist_download_excel(call):
        if not is_admin(call.from_user.id):
            bot.answer_callback_query(call.id, "🚫"); return
        rest = call.data[len("hist_excel_"):]; last_us = rest.rfind("_")
        date_str = rest[:last_us]; para = int(rest[last_us + 1:])
        with db_cursor() as (_, cursor):
            cursor.execute(
                "SELECT student_name,status FROM attendance "
                "WHERE date=%s AND para=%s ORDER BY student_name", (date_str, para))
            records = cursor.fetchall()
            cursor.execute(
                "SELECT DISTINCT subject FROM attendance WHERE date=%s AND para=%s",
                (date_str, para))
            sr = cursor.fetchone()
        subject = sr[0] if sr else "—"
        path = generate_attendance_excel(records, None, date_str, para, subject)
        if path:
            ok = send_excel_file(bot, call.message.chat.id, path,
                caption=f"📊 Барлау: {date_str} | {para}-пара: {subject}")
            bot.answer_callback_query(
                call.id, "✅ Excel жиберилди!" if ok else "❌ Жибериу қатеси.")
        else:
            bot.answer_callback_query(call.id, "❌ Excel ислмеды.")
